"""主窗口 - 完整实现设计稿"""
import json
import os
from datetime import datetime
from pathlib import Path

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QLineEdit, QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QFrame, QSizePolicy, QScrollArea, QMessageBox, QFileDialog, QCheckBox,
    QApplication, QStackedWidget, QGridLayout, QSpacerItem, QMenu, QComboBox,
    QProgressBar, QToolButton, QGraphicsOpacityEffect, QStyledItemDelegate
)
from PyQt6.QtCore import Qt, QTimer, pyqtSignal, QPoint, QSize, QEvent, QPropertyAnimation, QEasingCurve
from PyQt6.QtGui import QFont, QColor, QPalette, QAction, QClipboard

from models import Asset
from database import Database
from utils import (
    format_now, mask_serial, check_password_strength, calculate_status,
    calculate_alerts, STATUS_CONFIG, ALERT_CONFIG, TYPE_CONFIG, get_type_config
)
from templates import (
    LEGACY_TYPES, FIELD_PASSWORD, get_template, ordered_templates,
    resolve_column, TEMPLATE_GROUPS
)
from ui.style import MAIN_STYLE
from ui.dialogs import (
    AssetDialog, DeleteConfirmDialog, RecycleBinDialog,
    MasterPasswordDialog, ChangeMasterPasswordDialog, TemplateManagerDialog,
)


class StatCard(QFrame):
    """核心状态卡：可点击，点击后过滤下方表格（on_click 回调由主窗口注入）"""

    def __init__(self, label, value, color, top_color=None, sublabel="", parent=None):
        super().__init__(parent)
        self.setObjectName("statCard")
        self.setMinimumHeight(90)
        self.setMaximumHeight(100)
        self.on_click = None  # 点击回调
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setToolTip("点击过滤下方表格")
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 10, 12, 10)
        layout.setSpacing(4)

        if top_color:
            self.setStyleSheet(f"""
                #statCard {{
                    background: #ffffff;
                    border: 1px solid #e2e8f0;
                    border-radius: 10px;
                    border-top: 3px solid {top_color};
                }}
                #statCard:hover {{
                    border: 1px solid #c7d2fe;
                    border-top: 3px solid {top_color};
                }}
            """)
        else:
            self.setStyleSheet("""
                #statCard {
                    background: #ffffff;
                    border: 1px solid #e2e8f0;
                    border-radius: 10px;
                }
                #statCard:hover {
                    border: 1px solid #c7d2fe;
                }
            """)

        label_widget = QLabel(label)
        label_widget.setStyleSheet(f"font-size: 11px; font-weight: 600; color: {color if top_color else '#64748b'}; letter-spacing: 0.5px;")
        label_widget.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(label_widget)

        value_widget = QLabel(str(value))
        value_widget.setStyleSheet(f"font-size: 28px; font-weight: 700; color: {color}; line-height: 1;")
        value_widget.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(value_widget)

        if sublabel:
            sub_widget = QLabel(sublabel)
            sub_widget.setStyleSheet("font-size: 11px; color: #94a3b8; margin-top: 4px;")
            sub_widget.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(sub_widget)

        self.label_widget = label_widget
        self.value_label = value_widget

    def set_value(self, value):
        self.value_label.setText(str(value))

    def set_label(self, label):
        self.label_widget.setText(label)

    def mousePressEvent(self, event):
        if callable(self.on_click):
            self.on_click()
        super().mousePressEvent(event)


class RowHighlightDelegate(QStyledItemDelegate):
    """选中行高亮绘制。窗口级 + 控件级 QSS 同时接管 ::item 绘制时，
    模型的 BackgroundRole 会被忽略，选中底色必须由委托显式填充才能稳定可见。"""

    def __init__(self, is_selected, parent=None):
        super().__init__(parent)
        self._is_selected = is_selected  # 回调: (row) -> bool

    def paint(self, painter, option, index):
        if self._is_selected(index.row()):
            painter.save()
            painter.setPen(Qt.PenStyle.NoPen)
            painter.setBrush(QColor("#eef2ff"))
            painter.drawRect(option.rect)
            painter.restore()
        super().paint(painter, option, index)


class SidebarButton(QPushButton):
    def __init__(self, icon, text, count, active=False, active_style="type", color=None, parent=None):
        super().__init__(parent)
        self.icon = icon
        self.text_text = text
        self.count = count
        self.active = active
        self.active_style = active_style
        self.color = color  # 状态按钮的文字颜色，切换选中态后保持
        self.update_style()
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setMinimumHeight(36)
        self.setMaximumHeight(36)
        self.setText(f"{icon} {text}")

    def update_style(self):
        if self.active:
            if self.active_style == "type":
                # 选中态：左侧 4px 类型色竖条 + 浅灰背景，选中感更强
                bar_color = self.color or "#6366f1"
                text_color = self.color or "#0f172a"
                self.setStyleSheet(f"""
                    QPushButton {{
                        background: #f1f5f9; color: {text_color}; font-weight: 600;
                        border: none; border-left: 4px solid {bar_color};
                        border-radius: 8px;
                        padding: 8px 12px 8px 8px; margin: 1px 12px;
                        text-align: left; font-size: 13px;
                    }}
                    QPushButton:hover {{ background: #e2e8f0; }}
                """)
            else:
                bar_color = self.color or "#6366f1"
                text_color = self.color or "#4338ca"
                self.setStyleSheet(f"""
                    QPushButton {{
                        background: #eef2ff; color: {text_color}; font-weight: 600;
                        border: none; border-left: 4px solid {bar_color};
                        border-radius: 8px;
                        padding: 8px 12px 8px 8px; margin: 1px 12px;
                        text-align: left; font-size: 13px;
                    }}
                    QPushButton:hover {{ background: #dbeafe; }}
                """)
        else:
            text_color = self.color or "#475569"
            self.setStyleSheet(f"""
                QPushButton {{
                    background: transparent; color: {text_color}; font-weight: 500;
                    border: none; border-radius: 8px;
                    padding: 8px 12px; margin: 1px 12px;
                    text-align: left; font-size: 13px;
                }}
                QPushButton:hover {{ background: #f8fafc; }}
            """)

    def set_active(self, active):
        self.active = active
        self.update_style()


class MainWindow(QMainWindow):
    def __init__(self, db: Database):
        super().__init__()
        self.db = db
        self.assets = []
        self.recycle_bin = []
        self.selected_ids = set()
        self.current_type = "all"
        self.current_statuses = set()  # 状态筛选：多选，空集合 = 全部状态
        self.current_alert = None      # 告警筛选：单选，None = 不限（来自全局告警中心）
        self.hovered_row = -1          # 表格悬停行：-1 = 无（驱动复选框悬停显示）
        self.search_text = ""
        self.sort_field = "updated"
        self.sort_asc = False
        self.show_plaintext = True  # 账号/序列号默认明文显示，可在工具栏切换掩码
        self.custom_templates = self.db.get_custom_templates()  # 用户自定义资产模板

        # 完整描述仅保留在窗口标题栏；版本号见「设置 → 关于」，界面内不重复
        self.setWindowTitle("AssetVault | 个人数字资产管理器")
        # 最小高度适配 768p 笔记本屏幕（扣除任务栏后可用高度约 728px）
        self.setMinimumSize(1000, 560)
        self.resize(1400, 900)
        self.setup_ui()
        self.load_data()

    def setup_ui(self):
        self.setStyleSheet(MAIN_STYLE)
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # ===== 顶部品牌栏（极简单行：图标 + 产品名，高度 52px，版本号见「设置 → 关于」） =====
        header = QWidget()
        header.setObjectName("headerWidget")
        header.setMinimumHeight(52)
        header.setMaximumHeight(52)
        hlayout = QHBoxLayout(header)
        hlayout.setContentsMargins(20, 0, 20, 0)
        hlayout.setSpacing(10)

        logo = QLabel("🔐")
        logo.setObjectName("headerLogo")
        logo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        hlayout.addWidget(logo)

        title = QLabel("AssetVault")
        title.setObjectName("headerTitle")
        hlayout.addWidget(title)
        hlayout.addStretch()
        main_layout.addWidget(header)

        # ===== 主体 =====
        body = QWidget()
        body_layout = QHBoxLayout(body)
        body_layout.setContentsMargins(0, 0, 0, 0)
        body_layout.setSpacing(0)

        # ===== 左侧边栏 =====
        # 结构：筛选区（可滚动）+ 底部按钮（固定，永不因窗口缩小而消失）
        sidebar = QWidget()
        sidebar.setObjectName("sidebarWidget")
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(0, 16, 0, 16)
        sidebar_layout.setSpacing(0)

        # ----- 可滚动的筛选区 -----
        filter_widget = QWidget()
        filter_widget.setObjectName("sidebarFilterWidget")
        filter_layout = QVBoxLayout(filter_widget)
        filter_layout.setContentsMargins(0, 0, 0, 0)
        filter_layout.setSpacing(0)

        # 资产类型（按模板引擎动态生成：内置 + 自定义）
        type_title = QLabel("资产类型")
        type_title.setObjectName("sidebarSectionTitle")
        filter_layout.addWidget(type_title)

        self.type_buttons_container = QWidget()
        self.type_buttons_layout = QVBoxLayout(self.type_buttons_container)
        self.type_buttons_layout.setContentsMargins(0, 0, 0, 0)
        self.type_buttons_layout.setSpacing(0)
        filter_layout.addWidget(self.type_buttons_container)
        self.type_buttons = {}
        self._build_type_buttons()

        # 侧栏专注「资产类型」维度；状态过滤上移为表格上方的顶部标签栏
        filter_layout.addStretch()

        sidebar_scroll = QScrollArea()
        sidebar_scroll.setObjectName("sidebarScroll")
        sidebar_scroll.setWidgetResizable(True)
        sidebar_scroll.setFrameShape(QFrame.Shape.NoFrame)
        sidebar_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        sidebar_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        sidebar_scroll.setStyleSheet("""
            QScrollArea#sidebarScroll {
                background: transparent;
                border: none;
            }
            QScrollArea#sidebarScroll QScrollBar:vertical {
                background: transparent;
                width: 8px;
                margin: 4px 2px 4px 0;
            }
            QScrollArea#sidebarScroll QScrollBar::handle:vertical {
                background: #cbd5e1;
                border-radius: 4px;
                min-height: 30px;
            }
            QScrollArea#sidebarScroll QScrollBar::handle:vertical:hover {
                background: #94a3b8;
            }
            QScrollArea#sidebarScroll QScrollBar::add-line:vertical,
            QScrollArea#sidebarScroll QScrollBar::sub-line:vertical {
                height: 0;
            }
        """)
        sidebar_scroll.setWidget(filter_widget)
        sidebar_layout.addWidget(sidebar_scroll, 1)

        sidebar_layout.addSpacing(10)

        # 设置 + 回收站（并排排列，设置在左、回收站在右）
        btn_row_style = """
            QPushButton {
                background: #ffffff; color: #475569; font-weight: 500;
                border: 1px solid #e2e8f0; border-radius: 8px;
                padding: 8px 12px;
                font-size: 13px;
            }
            QPushButton:hover { background: #f8fafc; }
        """

        self.settings_btn = QPushButton("⚙️ 设置")
        self.settings_btn.setStyleSheet(btn_row_style)
        self.settings_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.settings_btn.clicked.connect(self.open_settings)

        recycle_btn = QPushButton("🗑️ 回收站")
        recycle_btn.setStyleSheet(btn_row_style)
        recycle_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        recycle_btn.clicked.connect(self.open_recycle_bin)

        self.recycle_count_label = QLabel("0")
        self.recycle_count_label.setStyleSheet("font-size: 11px; color: #ef4444; background: #fef2f2; border-radius: 10px; padding: 1px 7px; font-weight: 600;")

        bottom_layout = QHBoxLayout()
        bottom_layout.setContentsMargins(16, 0, 16, 0)
        bottom_layout.setSpacing(8)
        bottom_layout.addWidget(self.settings_btn, 1)
        bottom_layout.addWidget(recycle_btn, 1)
        bottom_layout.addWidget(self.recycle_count_label)
        sidebar_layout.addLayout(bottom_layout)

        # 侧栏固定宽度，直接加入主体（底部按钮已固定在侧栏内，无需外层滚动）
        sidebar.setMinimumWidth(240)
        sidebar.setMaximumWidth(240)
        body_layout.addWidget(sidebar)

        # ===== 右侧内容区 =====
        content = QScrollArea()
        content.setWidgetResizable(True)
        content.setFrameShape(QFrame.Shape.NoFrame)
        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(24, 24, 28, 24)
        content_layout.setSpacing(0)

        # 页面标题
        page_header = QWidget()
        ph_layout = QHBoxLayout(page_header)
        ph_layout.setContentsMargins(0, 0, 0, 0)
        ph_layout.setSpacing(0)
        page_title_layout = QVBoxLayout()
        page_title_layout.setSpacing(4)
        self.page_title = QLabel("全部资产")
        self.page_title.setStyleSheet("font-size: 22px; font-weight: 700; color: #0f172a; letter-spacing: -0.3px;")
        page_title_layout.addWidget(self.page_title)
        self.page_subtitle = QLabel("管理您的序列号、密码与数字资产")
        self.page_subtitle.setStyleSheet("font-size: 13px; color: #666666;")
        page_title_layout.addWidget(self.page_subtitle)
        ph_layout.addLayout(page_title_layout)
        ph_layout.addStretch()

        import_btn = QPushButton("📥 导入")
        import_btn.setObjectName("btnSecondary")
        import_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        import_btn.clicked.connect(self.import_data)
        ph_layout.addWidget(import_btn)

        export_btn = QPushButton("📤 导出")
        export_btn.setObjectName("btnSecondary")
        export_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        export_btn.setMenu(self.create_export_menu())
        ph_layout.addWidget(export_btn)

        self.add_btn = QPushButton("+ 新增资产")
        self.add_btn.setObjectName("btnPrimary")
        self.add_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.add_btn.setMenu(self.create_add_menu())
        ph_layout.addWidget(self.add_btn)
        content_layout.addWidget(page_header)
        content_layout.addSpacing(24)

        # ===== 统计面板：核心状态卡（始终显示，大字号） =====
        stats_row1 = QWidget()
        stats_row1_layout = QHBoxLayout(stats_row1)
        stats_row1_layout.setContentsMargins(0, 0, 0, 0)
        stats_row1_layout.setSpacing(16)

        self.stat_cards = {}
        configs = [
            ("total", "📋 资产总数", "0", "#0f172a", None, ""),
            ("normal", "正常", "0", "#22c55e", "#22c55e", ""),
            ("empty", "已用完", "0", "#ef4444", "#ef4444", ""),
            ("expired", "已过期", "0", "#f97316", "#f97316", ""),
        ]
        for key, label, value, color, top, sub in configs:
            card = StatCard(label, value, color, top, sub)
            # 卡片点击直接过滤表格：总数→全部状态，其余→对应状态标签
            card.on_click = lambda k=key: self.on_status_tab("all" if k == "total" else k)
            self.stat_cards[key] = card
            stats_row1_layout.addWidget(card)
        content_layout.addWidget(stats_row1)
        content_layout.addSpacing(10)

        # ===== 全局告警通知栏：即将到期 / 用量紧张 / 弱密码（可叠加标签，点击筛选表格） =====
        self.alert_bar = QWidget()
        self.alert_bar.setObjectName("alertBar")
        self.alert_bar.setStyleSheet("""
            #alertBar {
                background: #fffbeb;
                border: 1px solid #fde68a;
                border-radius: 10px;
            }
        """)
        alert_bar_layout = QHBoxLayout(self.alert_bar)
        alert_bar_layout.setContentsMargins(14, 8, 14, 8)
        alert_bar_layout.setSpacing(10)
        alert_title = QLabel("🔔 告警中心")
        alert_title.setStyleSheet("font-size: 13px; font-weight: 600; color: #92400e; background: transparent; border: none;")
        alert_bar_layout.addWidget(alert_title)
        self.alert_chips_layout = QHBoxLayout()
        self.alert_chips_layout.setContentsMargins(0, 0, 0, 0)
        self.alert_chips_layout.setSpacing(8)
        alert_bar_layout.addLayout(self.alert_chips_layout)
        alert_bar_layout.addStretch()
        self.alert_bar.setVisible(False)  # 无告警时整条隐藏，不占视觉空间
        content_layout.addWidget(self.alert_bar)
        content_layout.addSpacing(10)

        content_layout.addSpacing(10)

        # ===== 表格工具栏 =====
        toolbar = QWidget()
        toolbar_layout = QHBoxLayout(toolbar)
        toolbar_layout.setContentsMargins(0, 0, 0, 0)
        toolbar_layout.setSpacing(8)

        self.batch_actions_widget = QWidget()
        batch_layout = QHBoxLayout(self.batch_actions_widget)
        batch_layout.setContentsMargins(0, 0, 0, 0)
        batch_layout.setSpacing(8)

        self.selected_count_label = QLabel("已选 0 项")
        self.selected_count_label.setStyleSheet("font-size: 13px; color: #4338ca; font-weight: 600; padding: 4px 10px; background: #eef2ff; border-radius: 6px;")
        self.selected_count_label.setVisible(False)
        batch_layout.addWidget(self.selected_count_label)

        self.edit_btn = QPushButton("✏️ 编辑")
        self.edit_btn.setObjectName("btnSmall")
        self.edit_btn.setVisible(False)
        self.edit_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.edit_btn.clicked.connect(self.on_edit)
        batch_layout.addWidget(self.edit_btn)

        self.use_once_btn = QPushButton("📌 使用一次")
        self.use_once_btn.setObjectName("btnWarning")
        self.use_once_btn.setVisible(False)
        self.use_once_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.use_once_btn.clicked.connect(self.on_use_once)
        batch_layout.addWidget(self.use_once_btn)

        self.delete_btn = QPushButton("🗑️ 删除")
        self.delete_btn.setObjectName("btnDanger")
        self.delete_btn.setVisible(False)
        self.delete_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.delete_btn.clicked.connect(self.on_delete)
        batch_layout.addWidget(self.delete_btn)
        toolbar_layout.addWidget(self.batch_actions_widget)

        # 左侧：视图控制（明文/掩码切换，默认明文）
        self.plaintext_btn = QPushButton("👁 明文显示")
        self.plaintext_btn.setObjectName("btnSecondary")
        self.plaintext_btn.setCheckable(True)
        self.plaintext_btn.setChecked(True)
        self.plaintext_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.plaintext_btn.setToolTip("切换账号/序列号的明文与掩码显示")
        self.plaintext_btn.toggled.connect(self.on_toggle_plaintext)
        toolbar_layout.addWidget(self.plaintext_btn)

        # 中间：检索（弹性宽度）
        search_widget = QWidget()
        search_layout = QHBoxLayout(search_widget)
        search_layout.setContentsMargins(0, 0, 0, 0)
        search_layout.setSpacing(0)
        search_icon = QLabel("🔍")
        search_icon.setStyleSheet("font-size: 13px; color: #94a3b8; padding-left: 10px;")
        search_layout.addWidget(search_icon)
        self.search_input = QLineEdit()
        self.search_input.setObjectName("searchInput")
        self.search_input.setPlaceholderText("搜索名称、账号、邮箱...")
        self.search_input.textChanged.connect(self.on_search)
        search_layout.addWidget(self.search_input)
        toolbar_layout.addWidget(search_widget, 1)

        # 右侧：高级筛选开关（类型 / 有效期范围 / 标签组合筛选）
        self.adv_btn = QPushButton("🎯 高级筛选")
        self.adv_btn.setObjectName("btnSecondary")
        self.adv_btn.setCheckable(True)
        self.adv_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.adv_btn.setToolTip("按类型、有效期范围、标签组合筛选")
        self.adv_btn.toggled.connect(self.on_adv_toggled)
        toolbar_layout.addWidget(self.adv_btn)
        content_layout.addWidget(toolbar)
        content_layout.addSpacing(12)

        # ===== 高级筛选面板（默认收起） =====
        self.adv_panel = QWidget()
        self.adv_panel.setObjectName("advFilterPanel")
        self.adv_panel.setStyleSheet("""
            #advFilterPanel {
                background: #f8fafc;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
            }
            #advFilterPanel QLabel { background: transparent; font-size: 12px; color: #64748b; font-weight: 600; }
            #advFilterPanel QComboBox, #advFilterPanel QLineEdit {
                background: #ffffff; border: 1px solid #e2e8f0; border-radius: 8px;
                padding: 6px 10px; font-size: 13px; color: #334155; min-width: 130px;
            }
            #advFilterPanel QComboBox:hover, #advFilterPanel QLineEdit:focus { border-color: #6366f1; }
            #advFilterPanel QComboBox::drop-down { border: none; width: 20px; }
            #advFilterPanel QComboBox QAbstractItemView {
                background: #ffffff; border: 1px solid #e2e8f0; border-radius: 8px;
                padding: 4px; selection-background-color: #eef2ff; selection-color: #1e293b;
            }
        """)
        adv_layout = QHBoxLayout(self.adv_panel)
        adv_layout.setContentsMargins(14, 10, 14, 10)
        adv_layout.setSpacing(10)

        adv_layout.addWidget(QLabel("类型"))
        self.adv_type_combo = QComboBox()
        self.adv_type_combo.setCursor(Qt.CursorShape.PointingHandCursor)
        self.adv_type_combo.addItem("全部类型", None)
        for tpl in ordered_templates(self.custom_templates):
            self.adv_type_combo.addItem(f"{tpl['icon']} {tpl['label']}", tpl["key"])
        self.adv_type_combo.currentIndexChanged.connect(self.on_adv_changed)
        adv_layout.addWidget(self.adv_type_combo)

        adv_layout.addWidget(QLabel("有效期"))
        self.adv_expire_combo = QComboBox()
        self.adv_expire_combo.setCursor(Qt.CursorShape.PointingHandCursor)
        for label, key in [("全部有效期", None), ("7 天内到期", "7d"), ("30 天内到期", "30d"),
                           ("已过期", "expired"), ("永久", "permanent")]:
            self.adv_expire_combo.addItem(label, key)
        self.adv_expire_combo.currentIndexChanged.connect(self.on_adv_changed)
        adv_layout.addWidget(self.adv_expire_combo)

        adv_layout.addWidget(QLabel("标签"))
        self.adv_tags_input = QLineEdit()
        self.adv_tags_input.setPlaceholderText("多个标签用逗号分隔")
        self.adv_tags_input.textChanged.connect(self.on_adv_changed)
        adv_layout.addWidget(self.adv_tags_input, 1)

        adv_clear = QPushButton("✕ 清除")
        adv_clear.setObjectName("btnSmall")
        adv_clear.setCursor(Qt.CursorShape.PointingHandCursor)
        adv_clear.clicked.connect(self.clear_adv_filters)
        adv_layout.addWidget(adv_clear)

        self.adv_panel.setVisible(False)
        content_layout.addWidget(self.adv_panel)
        content_layout.addSpacing(12)
        # 高级筛选状态（None = 不限）
        self.adv_type = None
        self.adv_expire = None
        self.adv_tags = ""

        # ===== 状态标签栏（视图维度，与左侧类型导航物理分离） =====
        tabs_widget = QWidget()
        tabs_layout = QHBoxLayout(tabs_widget)
        tabs_layout.setContentsMargins(0, 0, 0, 0)
        tabs_layout.setSpacing(8)
        self.status_tabs = {}
        for key, label in [("all", "全部状态"), ("normal", "正常"), ("empty", "已用完"),
                           ("expired", "已过期"), ("archived", "已归档")]:
            btn = QPushButton(label)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.clicked.connect(lambda checked=False, k=key: self.on_status_tab(k))
            self.status_tabs[key] = btn
            tabs_layout.addWidget(btn)
        tabs_layout.addStretch()
        self._refresh_status_tabs()
        content_layout.addWidget(tabs_widget)
        content_layout.addSpacing(10)

        # ===== 数据表格（含空状态切换） =====
        self.table = QTableWidget()
        self.table.setObjectName("dataTable")
        self.table.setColumnCount(12)
        self.table.setHorizontalHeaderLabels([
            "", "名称", "类型", "账号/序列号", "绑定邮箱", "使用/数量",
            "有效期", "状态", "⚠ 告警", "备注", "更新时间", "操作"
        ])
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(48)
        header = self.table.horizontalHeader()
        # 列宽策略：固定窄列（状态/类型）、中等列（有效期/更新时间/操作）、
        # 弹性宽列（名称/备注），账号与邮箱可拖拽调整
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        header.setMinimumSectionSize(24)  # 允许复选框列窄于表头 padding 推导出的最小宽度
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(0, 32)  # 复选框列：仅悬停/选中时显示，列宽收窄
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(2, 100)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(3, 140)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(4, 110)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(5, 90)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(6, 110)  # 仅显示日期，剩余天数见 tooltip
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(7, 70)
        header.setSectionResizeMode(8, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(8, 80)  # 告警列固定 80px，禁止压缩，避免图标换行/截断
        header.setSectionResizeMode(10, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(10, 140)
        header.setSectionResizeMode(11, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(11, 110)
        # 表头对齐：文本左对齐，数字/状态/告警居中，日期右对齐
        header_align = {
            1: Qt.AlignmentFlag.AlignLeft, 2: Qt.AlignmentFlag.AlignCenter,
            3: Qt.AlignmentFlag.AlignLeft, 4: Qt.AlignmentFlag.AlignLeft,
            5: Qt.AlignmentFlag.AlignCenter, 6: Qt.AlignmentFlag.AlignRight,
            7: Qt.AlignmentFlag.AlignCenter, 8: Qt.AlignmentFlag.AlignCenter,
            9: Qt.AlignmentFlag.AlignLeft, 10: Qt.AlignmentFlag.AlignRight,
            11: Qt.AlignmentFlag.AlignCenter,
        }
        for col, ali in header_align.items():
            item = self.table.horizontalHeaderItem(col)
            if item:
                item.setTextAlignment(ali | Qt.AlignmentFlag.AlignVCenter)
        self.table.setAlternatingRowColors(False)
        self.table.setShowGrid(False)
        self.table.setStyleSheet("""
            QTableWidget {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
                gridline-color: transparent;
            }
            QTableWidget::item {
                padding: 10px 10px;
                border-bottom: 1px solid #f1f5f9;
                color: #475569;
                font-size: 13px;
            }
            QTableWidget::item:hover {
                background: #f8fafc;
            }
            QTableWidget::item:selected {
                background: #eef2ff;
                color: #1e293b;
            }
            QHeaderView::section {
                background: #f8fafc;
                color: #475569;
                font-size: 12px;
                font-weight: 600;
                letter-spacing: 0.3px;
                padding: 12px 10px;
                border: none;
                border-bottom: 1px solid #e2e8f0;
            }
            QHeaderView::section:hover {
                background: #f1f5f9;
            }
        """)
        self.table.horizontalHeader().sectionClicked.connect(self.on_header_clicked)
        # 右键菜单：通过事件过滤器捕获鼠标右键抬起（跨平台稳定，
        # 不依赖各平台对 ContextMenu 事件的合成差异）
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.PreventContextMenu)
        self.table.viewport().installEventFilter(self)
        # 双击行任意位置编辑该记录
        self.table.cellDoubleClicked.connect(self.on_row_double_clicked)
        # 复选框交互：悬停行显示复选框，单击行切换选中
        # 选中行高亮委托（与复选框联动，背景 #EEF2FF；不依赖 QSS 对模型背景色的支持）
        self.table.setItemDelegate(RowHighlightDelegate(
            lambda r: 0 <= r < len(getattr(self, "_table_assets", []))
                      and self._table_assets[r].id in self.selected_ids,
            self.table,
        ))
        self.table.setMouseTracking(True)
        self.table.viewport().setMouseTracking(True)
        self.table.cellEntered.connect(self.on_cell_entered)
        self.table.cellClicked.connect(self.on_cell_clicked)

        # ===== 空状态（表格无数据时替代显示，避免「白茫茫大地真干净」） =====
        empty_widget = QWidget()
        empty_widget.setObjectName("emptyStateWidget")
        empty_widget.setStyleSheet("""
            #emptyStateWidget {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
            }
        """)
        empty_widget.setMinimumHeight(320)
        el = QVBoxLayout(empty_widget)
        el.setContentsMargins(24, 40, 24, 40)
        el.setSpacing(0)
        el.addStretch()
        self.empty_icon = QLabel("📭")
        self.empty_icon.setObjectName("emptyIcon")
        self.empty_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        el.addWidget(self.empty_icon, 0, Qt.AlignmentFlag.AlignCenter)
        el.addSpacing(16)
        self.empty_title = QLabel("暂无资产")
        self.empty_title.setObjectName("emptyTitle")
        self.empty_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        el.addWidget(self.empty_title)
        el.addSpacing(8)
        self.empty_desc = QLabel("点击右上角「新增资产」开始录入")
        self.empty_desc.setObjectName("emptyDesc")
        self.empty_desc.setAlignment(Qt.AlignmentFlag.AlignCenter)
        el.addWidget(self.empty_desc)
        el.addSpacing(20)
        empty_btns = QHBoxLayout()
        empty_btns.setSpacing(10)
        empty_btns.addStretch()
        self.empty_clear_btn = QPushButton("清除筛选")
        self.empty_clear_btn.setObjectName("btnSecondary")
        self.empty_clear_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.empty_clear_btn.clicked.connect(self.clear_filters)
        empty_btns.addWidget(self.empty_clear_btn)
        self.empty_add_btn = QPushButton("+ 新增资产")
        self.empty_add_btn.setObjectName("btnPrimary")
        self.empty_add_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.empty_add_btn.clicked.connect(self.on_empty_add_clicked)
        empty_btns.addWidget(self.empty_add_btn)
        empty_btns.addStretch()
        el.addLayout(empty_btns)
        el.addStretch()

        self.table_stack = QStackedWidget()
        self.table_stack.addWidget(self.table)         # 页 0：数据表格
        self.table_stack.addWidget(empty_widget)       # 页 1：空状态
        content_layout.addWidget(self.table_stack)

        # 底部
        footer = QWidget()
        footer_layout = QHBoxLayout(footer)
        footer_layout.setContentsMargins(0, 12, 0, 0)
        footer_layout.setSpacing(0)
        footer_left = QLabel("🔒 数据文件: vault.db ｜ 本地 AES-256 加密存储")
        footer_left.setStyleSheet("font-size: 12px; color: #94a3b8;")
        footer_right = QLabel("AssetVault v2.0")
        footer_right.setStyleSheet("font-size: 12px; color: #94a3b8;")
        footer_layout.addWidget(footer_left)
        footer_layout.addStretch()
        footer_layout.addWidget(footer_right)
        content_layout.addWidget(footer)

        content.setWidget(content_widget)
        body_layout.addWidget(content)
        main_layout.addWidget(body)

        # Toast
        self.toast = QLabel("")
        self.toast.setStyleSheet("""
            background: #0f172a; color: white;
            border-radius: 10px; padding: 12px 20px;
            font-size: 13px; font-weight: 500;
        """)
        self.toast.setVisible(False)
        self.toast.setParent(self)
        self.toast.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.WindowStaysOnTopHint)
        self.toast.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)

    # ==================== 状态标签栏 & 高级筛选 ====================

    def on_status_tab(self, key):
        """状态标签栏：单选视图过滤（全部状态 = 不限），与左侧类型导航物理分离"""
        self.current_statuses = set() if key == "all" else {key}
        self._refresh_status_tabs()
        self.selected_ids.clear()
        self.update_table()

    def _refresh_status_tabs(self):
        """刷新状态标签选中态样式（药丸样式，选中按状态色高亮）"""
        style_map = {
            "all": ("#6366f1", "#eef2ff"), "normal": ("#22c55e", "#f0fdf4"),
            "empty": ("#ef4444", "#fef2f2"), "expired": ("#f97316", "#fff7ed"),
            "archived": ("#64748b", "#f8fafc"),
        }
        active = "all" if not self.current_statuses else next(iter(self.current_statuses))
        for key, btn in self.status_tabs.items():
            color, bg = style_map[key]
            if key == active:
                btn.setStyleSheet(
                    f"QPushButton {{ background: {bg}; color: {color}; border: 1px solid {color}; "
                    f"border-radius: 14px; padding: 5px 16px; font-size: 12px; font-weight: 600; }}")
            else:
                btn.setStyleSheet(
                    "QPushButton { background: transparent; color: #64748b; border: 1px solid #e2e8f0; "
                    "border-radius: 14px; padding: 5px 16px; font-size: 12px; }"
                    "QPushButton:hover { background: #f1f5f9; }")

    def _update_status_tab_counts(self, context_assets):
        """状态标签计数跟随上下文（选中类型时只统计该类型）；已归档不计入活跃资产"""
        counts = {"all": 0, "normal": 0, "empty": 0, "expired": 0, "archived": 0}
        for a in context_assets:
            if a.status == "archived":
                counts["archived"] += 1
            elif a.status in counts:
                counts[a.status] += 1
                counts["all"] += 1  # 全部状态 = 活跃资产总数（不含已归档）
        labels = {"all": "全部状态", "normal": "正常", "empty": "已用完",
                  "expired": "已过期", "archived": "已归档"}
        for key, btn in self.status_tabs.items():
            btn.setText(f"{labels[key]} {counts[key]}")

    def on_adv_toggled(self, checked):
        """高级筛选面板开关；关闭时自动清除高级筛选条件"""
        self.adv_panel.setVisible(checked)
        if not checked:
            self.clear_adv_filters()

    def on_adv_changed(self, *args):
        self.adv_type = self.adv_type_combo.currentData()
        self.adv_expire = self.adv_expire_combo.currentData()
        self.adv_tags = self.adv_tags_input.text().strip()
        self.update_table()

    def clear_adv_filters(self):
        for w in (self.adv_type_combo, self.adv_expire_combo, self.adv_tags_input):
            w.blockSignals(True)
        self.adv_type_combo.setCurrentIndex(0)
        self.adv_expire_combo.setCurrentIndex(0)
        self.adv_tags_input.clear()
        for w in (self.adv_type_combo, self.adv_expire_combo, self.adv_tags_input):
            w.blockSignals(False)
        self.adv_type = None
        self.adv_expire = None
        self.adv_tags = ""
        self.update_table()

    def create_add_menu(self):
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
                padding: 4px;
            }
            QMenu::item {
                padding: 10px 14px;
                font-size: 13px;
                color: #475569;
                border-radius: 6px;
            }
            QMenu::item:selected {
                background: #f8fafc;
            }
            QMenu::separator {
                height: 1px;
                background: #f1f5f9;
                margin: 4px 10px;
            }
        """)
        # 全部模板类型（内置 + 自定义）
        for tpl in ordered_templates(self.custom_templates):
            action = QAction(f"{tpl['icon']} 新增{tpl['label']}", self)
            action.triggered.connect(lambda checked=False, k=tpl["key"]: self.open_add_dialog(k))
            menu.addAction(action)
        return menu

    def show_toast(self, msg):
        self.toast.setText(msg)
        self.toast.adjustSize()
        self.toast.move(
            self.width() - self.toast.width() - 24,
            self.height() - self.toast.height() - 24
        )
        self.toast.setVisible(True)
        QTimer.singleShot(2500, lambda: self.toast.setVisible(False))

    def load_data(self):
        self.custom_templates = self.db.get_custom_templates()
        self.assets = self.db.get_assets()
        self.recycle_bin = self.db.get_recycle_bin()
        self.update_stats()
        self.update_table()

    def _context_assets(self):
        """当前上下文资产集：选中具体类型时只统计该类型"""
        if self.current_type == "all":
            return list(self.assets)
        return [a for a in self.assets if a.asset_type == self.current_type]

    def update_stats(self):
        """核心状态卡（生命周期，上下文感知）+ 全局告警栏 + 侧栏计数"""
        ctx = self._context_assets()
        active = [a for a in ctx if a.status != "archived"]  # 已归档不计入活跃资产
        values = {
            "total": len(active),
            "normal": sum(1 for a in active if a.status == "normal"),
            "empty": sum(1 for a in active if a.status == "empty"),
            "expired": sum(1 for a in active if a.status == "expired"),
        }
        for key, card in self.stat_cards.items():
            card.set_value(values.get(key, 0))
        # 总数卡标签跟随上下文（如「密码总数」）
        if self.current_type == "all":
            self.stat_cards["total"].set_label("📋 资产总数")
        else:
            tc = get_type_config(self.current_type, self.custom_templates)
            self.stat_cards["total"].set_label(f"{tc['icon']} {tc['label']}总数")

        self._update_alert_bar(active)
        self.recycle_count_label.setText(str(len(self.recycle_bin)))

        # 侧边栏类型计数（活跃资产，实时更新）
        from collections import Counter
        active_all = [a for a in self.assets if a.status != "archived"]
        type_counts = Counter(a.asset_type for a in active_all)
        total = len(active_all)
        for key, btn in self.type_buttons.items():
            count = total if key == "all" else type_counts.get(key, 0)
            btn.setText(f"{btn.icon} {btn.text_text}    {count}")
        # 状态计数显示在顶部标签栏上（跟随上下文）
        self._update_status_tab_counts(ctx)

    # ==================== 全局告警中心 ====================

    def _asset_alerts(self, asset):
        """资产的健康度告警列表（按资产模板的到期提醒阈值计算）"""
        tpl = get_template(asset.asset_type, self.custom_templates)
        return calculate_alerts(asset, expire_days=tpl.get("expire_days"))

    def _update_alert_bar(self, active_assets):
        """刷新全局告警通知栏：统计各告警数量，无告警时整条隐藏"""
        # 清空旧 chips（脱离布局后手动删除，避免残留）
        while self.alert_chips_layout.count():
            item = self.alert_chips_layout.takeAt(0)
            w = item.widget()
            if w is not None:
                w.hide()
                w.setParent(None)
                w.deleteLater()
        counts = {}
        for a in active_assets:
            for key in self._asset_alerts(a):
                counts[key] = counts.get(key, 0) + 1
        if not counts:
            self.alert_bar.setVisible(False)
            # 上下文切换后当前告警筛选可能已无意义，自动清除
            if self.current_alert:
                self.current_alert = None
                self.update_table()
            return
        for key in ("expiring", "tight", "weak_password"):
            if key not in counts:
                continue
            cfg = ALERT_CONFIG[key]
            chip = QPushButton(f"{cfg['icon']} {cfg['label']} {counts[key]}")
            chip.setCursor(Qt.CursorShape.PointingHandCursor)
            chip.setToolTip(f"点击筛选出所有「{cfg['label']}」资产")
            chip.setProperty("alert_key", key)
            chip.clicked.connect(lambda checked=False, k=key: self.on_alert_chip(k))
            self.alert_chips_layout.addWidget(chip)
        self._refresh_alert_chips()
        self.alert_bar.setVisible(True)

    def _refresh_alert_chips(self):
        """告警 chip 选中态：选中的告警用实色边框高亮"""
        for i in range(self.alert_chips_layout.count()):
            chip = self.alert_chips_layout.itemAt(i).widget()
            if chip is None:
                continue
            key = chip.property("alert_key")
            cfg = ALERT_CONFIG[key]
            if key == self.current_alert:
                chip.setStyleSheet(
                    f"QPushButton {{ background: {cfg['bg']}; color: {cfg['color']}; "
                    f"border: 2px solid {cfg['color']}; border-radius: 12px; "
                    f"padding: 3px 12px; font-size: 12px; font-weight: 700; }}")
            else:
                chip.setStyleSheet(
                    f"QPushButton {{ background: {cfg['bg']}; color: {cfg['color']}; "
                    f"border: 1px solid {cfg['border']}; border-radius: 12px; "
                    f"padding: 3px 12px; font-size: 12px; font-weight: 600; }}"
                    f"QPushButton:hover {{ border: 1px solid {cfg['color']}; }}")

    def on_alert_chip(self, key):
        """点击告警 chip：筛选出带该告警的资产；再次点击取消筛选"""
        self.current_alert = None if self.current_alert == key else key
        self._refresh_alert_chips()
        self.selected_ids.clear()
        self.update_table()

    def get_filtered_assets(self):
        result = self.assets[:]
        if self.current_type != "all":
            result = [a for a in result if a.asset_type == self.current_type]
        if self.current_statuses:
            result = [a for a in result if a.status in self.current_statuses]
        else:
            result = [a for a in result if a.status != "archived"]  # 默认视图只显示活跃资产
        # 告警筛选（来自全局告警中心，与状态筛选叠加取交集）
        if self.current_alert:
            result = [a for a in result if self.current_alert in self._asset_alerts(a)]
        # 高级筛选：类型 / 有效期范围 / 标签组合（与左侧类型导航叠加取交集）
        if self.adv_type:
            result = [a for a in result if a.asset_type == self.adv_type]
        if self.adv_expire:
            if self.adv_expire == "permanent":
                result = [a for a in result if not a.expire]
            elif self.adv_expire == "expired":
                result = [a for a in result if (d := self._expire_days(a.expire)) is not None and d < 0]
            else:
                limit = int(self.adv_expire.rstrip("d"))
                result = [a for a in result if (d := self._expire_days(a.expire)) is not None and 0 <= d <= limit]
        if self.adv_tags:
            # 多个标签逗号分隔，AND 匹配（每个标签都命中才保留），大小写不敏感
            wanted = [t.strip().lower() for t in self.adv_tags.split(",") if t.strip()]
            for t in wanted:
                result = [a for a in result if t in (a.tags or "").lower()]
        if self.search_text:
            search = self.search_text.lower()
            result = [a for a in result if search in (a.name + a.account + (a.email or "") + (a.note or "") + (a.url or "") + (a.tags or "")).lower()]
        return result

    @staticmethod
    def _expire_days(expire):
        """有效期距今天数（None 表示无有效期或解析失败）"""
        if not expire:
            return None
        try:
            return (datetime.strptime(expire, "%Y-%m-%d").date() - datetime.now().date()).days
        except (ValueError, TypeError):
            return None

    def _expire_display(self, expire):
        """有效期单元格：(显示文本, 颜色)。仅显示日期；永久→「-」；保留临期变色提醒
        （已过期/≤7 天红，≤30 天橙），tooltip 补充剩余天数详情"""
        days = self._expire_days(expire)
        if days is None:
            if expire:  # 有值但解析失败，原样展示
                return expire, "#475569"
            return "-", "#cbd5e1"
        if days < 0 or days <= 7:
            return expire, "#ef4444"
        if days <= 30:
            return expire, "#f59e0b"
        return expire, "#475569"

    def _expire_tooltip(self, expire):
        """有效期 tooltip：剩余/已过期天数详情（单元格本身只显示日期）"""
        days = self._expire_days(expire)
        if days is None:
            return None
        if days < 0:
            return f"已过期 {-days} 天"
        return f"剩余 {days} 天"

    # 告警列图标方案：颜色圆点 + 16px 图标，横向排列（Tooltip 兜底完整含义）
    ALERT_ICONS = {
        "tight": "🟠⚡",           # 用量紧张：橙点 + 闪电
        "expiring": "🔴⏰",        # 即将到期：红点 + 时钟
        "weak_password": "🟡🛡️",   # 弱密码：黄点 + 盾牌
    }
    ALERT_MAX_ICONS = 3  # 单条资产告警超过 3 个时折叠为「+N」

    def _alert_tooltip(self, key, asset):
        """告警 Tooltip：完整告警文案 + 具体数据（如「剩余 10 天到期」「剩余 1 次可用」）"""
        if key == "tight":
            return f"用量紧张：剩余 {asset.remain or 0} 次可用（共 {asset.total or 0} 次）"
        if key == "expiring":
            days = self._expire_days(asset.expire)
            if days is not None:
                return f"即将到期：剩余 {days} 天到期（{asset.expire}）"
            return f"即将到期：{asset.expire}"
        if key == "weak_password":
            return "弱密码：密码强度低，建议更换"
        return ALERT_CONFIG[key]["label"]

    def _alert_badges_widget(self, asset, row=None, expanded=False):
        """告警列单元格：圆点+16px 图标横向排列（4px 间距），无告警显示 '-'

        告警超过 3 个时只显示前 3 个图标 + 「+N」灰色小字，点击展开全部。
        """
        alerts = self._asset_alerts(asset)
        container = QWidget()
        container.setStyleSheet("background: transparent;")
        layout = QHBoxLayout(container)
        layout.setContentsMargins(6, 0, 6, 0)
        layout.setSpacing(4)  # 图标横向排列，4px 间距
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        if not alerts:
            dash = QLabel("-")
            dash.setStyleSheet("color: #cbd5e1; background: transparent;")
            layout.addWidget(dash)
            return container
        shown = alerts if (expanded or len(alerts) <= self.ALERT_MAX_ICONS) else alerts[:self.ALERT_MAX_ICONS]
        for key in shown:
            icon = QLabel(self.ALERT_ICONS[key])
            icon.setToolTip(self._alert_tooltip(key, asset))  # 悬停显示完整告警文案
            icon.setStyleSheet("background: transparent; font-size: 16px;")
            layout.addWidget(icon)
        hidden = len(alerts) - len(shown)
        if hidden > 0:
            more = QPushButton(f"+{hidden}")
            more.setCursor(Qt.CursorShape.PointingHandCursor)
            # 悬停 +N 列出全部被折叠的告警
            more.setToolTip("\n".join(self._alert_tooltip(k, asset) for k in alerts))
            more.setStyleSheet("""
                QPushButton {
                    background: transparent; border: none;
                    color: #94a3b8; font-size: 11px; font-weight: 600; padding: 0 2px;
                }
                QPushButton:hover { color: #475569; }
            """)
            if row is not None:
                more.clicked.connect(lambda checked=False, r=row, a=asset: self._expand_alert_cell(r, a))
            layout.addWidget(more)
        return container

    def _expand_alert_cell(self, row, asset):
        """点击「+N」展开该单元格的全部告警图标"""
        old = self.table.cellWidget(row, 8)
        self.table.setCellWidget(row, 8, self._alert_badges_widget(asset, row=row, expanded=True))
        if old is not None:
            old.hide()
            old.setParent(None)
            old.deleteLater()

    def update_table(self):
        filtered = self.get_filtered_assets()
        self._table_assets = filtered  # 供右键菜单按行号取记录
        self.hovered_row = -1  # 行号随重建失效，悬停态待鼠标重新进入
        # 显式销毁旧的单元格控件（复选框/用量条/告警/操作列）：removeCellWidget 只脱离索引，
        # 控件仍以视口为父且保持可见，必须手动移除并删除，否则层层残留
        for r in range(self.table.rowCount()):
            for c in (0, 5, 8, 11):
                w = self.table.cellWidget(r, c)
                if w is not None:
                    self.table.removeCellWidget(r, c)
                    w.hide()
                    w.setParent(None)
                    w.deleteLater()
        self.table.setRowCount(len(filtered))
        self._update_empty_state(filtered)

        for i, asset in enumerate(filtered):
            st = STATUS_CONFIG.get(asset.status, STATUS_CONFIG["normal"])
            tc = get_type_config(asset.asset_type, self.custom_templates)
            is_selected = asset.id in self.selected_ids

            # 复选框
            checkbox = QCheckBox()
            checkbox.setChecked(is_selected)
            checkbox.setStyleSheet("""
                QCheckBox::indicator {
                    width: 16px; height: 16px;
                    border-radius: 4px; border: 1px solid #e2e8f0;
                    background: #ffffff;
                }
                QCheckBox::indicator:checked {
                    background: #6366f1; border-color: #6366f1;
                }
            """)
            checkbox.stateChanged.connect(lambda state, aid=asset.id: self.on_checkbox_changed(aid, state))
            checkbox_widget = QWidget()
            checkbox_layout = QHBoxLayout(checkbox_widget)
            checkbox_layout.setContentsMargins(0, 0, 0, 0)
            checkbox_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            checkbox_layout.addWidget(checkbox)
            self.table.setCellWidget(i, 0, checkbox_widget)
            # 默认隐藏复选框：仅已选中行常显，其余行悬停时显示。
            # 隐藏的是内部 QCheckBox 而非容器——Qt 视图会强制显示单元格控件本身，
            # 但子控件的可见性由我们完全控制（容器透明，隐藏子控件即视觉隐藏）
            checkbox.setVisible(is_selected)

            # 名称（继承应用字体设置粗体，避免空族名回退导致合成加粗发虚）
            name_item = QTableWidgetItem(asset.name)
            bold_font = QFont(QApplication.font())
            bold_font.setBold(True)
            name_item.setFont(bold_font)
            name_item.setToolTip(asset.name)  # 列宽不足省略时悬停查看全名
            self.table.setItem(i, 1, name_item)

            # 类型
            type_item = QTableWidgetItem(f"{tc['icon']} {tc['label']}")
            type_item.setForeground(QColor(tc["color"]))
            self.table.setItem(i, 2, type_item)

            # 账号/序列号：纯文本等宽显示，悬停 tooltip 查看完整内容（复制走右键菜单/操作列）
            account_display = asset.account if self.show_plaintext else mask_serial(asset.account)
            account_item = QTableWidgetItem(account_display or "-")
            account_item.setFont(QFont("Courier New", 10))
            account_item.setToolTip(account_display or "-")
            self.table.setItem(i, 3, account_item)

            # 绑定邮箱
            self.table.setItem(i, 4, QTableWidgetItem(asset.email or "-"))

            # 使用/数量：进度条 + 文字，颜色随剩余量（0 红 / ≤1 橙 / 其余绿）
            if asset.asset_type == "serial":
                total = asset.total or 0
                used = asset.used or 0
                remain = asset.remain or 0
                bar_color = "#ef4444" if remain == 0 else ("#f59e0b" if remain <= 1 else "#22c55e")
                usage_widget = QWidget()
                usage_widget.setProperty("cell_proxy", True)
                usage_widget.setStyleSheet("background: transparent;")
                usage_layout = QHBoxLayout(usage_widget)
                usage_layout.setContentsMargins(6, 0, 6, 0)
                usage_layout.setSpacing(6)
                bar = QProgressBar()
                bar.setRange(0, max(total, 1))
                bar.setValue(min(used, total) if total > 0 else 0)
                bar.setFormat(f"{used}/{total}")
                bar.setTextVisible(True)
                bar.setFixedHeight(16)
                bar.setStyleSheet(f"""
                    QProgressBar {{
                        background: #f1f5f9; border: none; border-radius: 8px;
                        font-size: 10px; color: #475569; text-align: center;
                    }}
                    QProgressBar::chunk {{ background: {bar_color}; border-radius: 8px; }}
                """)
                bar.setToolTip(f"已使用 {used} / 共 {total}，剩余 {remain}")
                usage_layout.addWidget(bar)
                self.table.setCellWidget(i, 5, usage_widget)
            else:
                usage_item = QTableWidgetItem("-")
                usage_item.setForeground(QColor("#94a3b8"))
                self.table.setItem(i, 5, usage_item)

            # 有效期：永久显示为「-」；有日期显示「日期 · 剩余 X 天」，临近到期自动变色
            expire_text, expire_color = self._expire_display(asset.expire)
            expire_item = QTableWidgetItem(expire_text)
            expire_item.setForeground(QColor(expire_color))
            tip = self._expire_tooltip(asset.expire)
            if tip:
                expire_item.setToolTip(f"{asset.expire}（{tip}）")
            if asset.expire:
                days = self._expire_days(asset.expire)
                if days is not None and days <= 30:
                    expire_item.setFont(bold_font)
            self.table.setItem(i, 6, expire_item)

            # 状态（仅生命周期：正常/已用完/已过期/已归档，用颜色区分）
            status_item = QTableWidgetItem(st["label"])
            status_item.setForeground(QColor(st["color"]))
            self.table.setItem(i, 7, status_item)

            # 告警（健康度徽章，可叠加，如「🔴 即将到期（3天）」；无告警显示 -）
            self.table.setCellWidget(i, 8, self._alert_badges_widget(asset, row=i))

            # 备注
            note_item = QTableWidgetItem(asset.note or "-")
            note_item.setForeground(QColor("#94a3b8"))
            note_item.setFont(QFont("", 10))
            self.table.setItem(i, 9, note_item)

            # 更新时间
            self.table.setItem(i, 10, QTableWidgetItem(asset.updated))

            # 操作列：编辑 / 复制主字段 / 归档（或恢复）/ 删除（固定列宽，始终在表格最右侧）
            ops_widget = QWidget()
            ops_widget.setStyleSheet("background: transparent;")
            ops_layout = QHBoxLayout(ops_widget)
            ops_layout.setContentsMargins(2, 0, 2, 0)
            ops_layout.setSpacing(2)
            ops_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            ops_btn_style = """
                QPushButton {
                    background: transparent; border: none; border-radius: 6px;
                    font-size: 13px; padding: 4px;
                }
                QPushButton:hover { background: #eef2ff; }
            """
            is_archived = asset.status == "archived"
            # 复制按钮提示语跟随类型：密码行复制密码，序列号行复制序列号，其余复制主字段
            copy_tip = {"serial": "复制序列号", "password": "复制密码"}.get(asset.asset_type, "复制账号/密码")
            for icon, tip, handler in [
                ("✏️", "编辑", lambda checked=False, aid=asset.id: self.on_edit_asset(aid)),
                ("📋", copy_tip, lambda checked=False, aid=asset.id: self.on_copy_asset(aid)),
                ("↩️" if is_archived else "📦",
                 "取消归档" if is_archived else "归档（手动停用，不计入活跃资产）",
                 lambda checked=False, aid=asset.id: self.on_toggle_archive(aid)),
                ("🗑️", "删除", lambda checked=False, aid=asset.id: self.on_delete_asset(aid)),
            ]:
                b = QPushButton(icon)
                b.setToolTip(tip)
                b.setFixedSize(28, 28)
                b.setCursor(Qt.CursorShape.PointingHandCursor)
                b.setStyleSheet(ops_btn_style)
                b.clicked.connect(handler)
                ops_layout.addWidget(b)
            self.table.setCellWidget(i, 11, ops_widget)

            # 单元格对齐：文本左对齐，数字/状态居中，日期右对齐
            cell_align = {
                1: Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
                2: Qt.AlignmentFlag.AlignCenter,
                3: Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
                4: Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
                5: Qt.AlignmentFlag.AlignCenter,
                6: Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter,
                7: Qt.AlignmentFlag.AlignCenter,
                9: Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
                10: Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter,
            }
            for col, ali in cell_align.items():
                item = self.table.item(i, col)
                if item:
                    item.setTextAlignment(ali)

            # 行背景
            if is_selected:
                for col in range(12):
                    item = self.table.item(i, col)
                    if item:
                        item.setBackground(QColor("#eef2ff"))

        self.update_batch_actions()

    def _update_empty_state(self, filtered):
        """空状态：区分「没有数据」与「筛选无结果」，并提供直达操作"""
        if filtered:
            self.table_stack.setCurrentIndex(0)
            return
        has_filter = (bool(self.search_text) or bool(self.current_statuses)
                      or bool(self.current_alert)
                      or bool(self.adv_type) or bool(self.adv_expire) or bool(self.adv_tags))
        if has_filter:
            self.empty_icon.setText("🔍")
            self.empty_title.setText("未找到匹配的资产")
            self.empty_desc.setText("尝试调整搜索关键词或筛选条件")
            self.empty_clear_btn.setVisible(True)
            self.empty_add_btn.setVisible(False)
        else:
            self.empty_icon.setText("📭")
            if self.current_type == "all":
                self.empty_title.setText("暂无资产")
                self.empty_desc.setText("点击右上角「新增资产」开始录入")
                self.empty_add_btn.setText("+ 新增资产")
            else:
                tc = get_type_config(self.current_type, self.custom_templates)
                self.empty_title.setText(f"暂无{tc['label']}资产")
                self.empty_desc.setText("点击右上角「新增资产」开始录入")
                self.empty_add_btn.setText(f"+ 新增{tc['label']}")
            self.empty_clear_btn.setVisible(False)
            self.empty_add_btn.setVisible(True)
        self.table_stack.setCurrentIndex(1)

    def on_empty_add_clicked(self):
        """空状态直达新增：全部视图弹出类型菜单，具体类型直接打开对应表单"""
        if self.current_type == "all":
            menu = self.create_add_menu()
            menu.exec(self.empty_add_btn.mapToGlobal(QPoint(0, self.empty_add_btn.height() + 4)))
        else:
            self.open_add_dialog(self.current_type)

    def clear_filters(self):
        """清除搜索、状态标签与高级筛选，恢复完整列表"""
        self.search_text = ""
        self.search_input.blockSignals(True)
        self.search_input.clear()
        self.search_input.blockSignals(False)
        self.current_statuses.clear()
        self._refresh_status_tabs()
        self.current_alert = None
        self._refresh_alert_chips()
        self.clear_adv_filters()
        self.selected_ids.clear()
        self.update_table()

    def update_batch_actions(self):
        count = len(self.selected_ids)
        self.selected_count_label.setVisible(count > 0)
        self.selected_count_label.setText(f"已选 {count} 项")
        self.edit_btn.setVisible(count == 1)
        self.use_once_btn.setVisible(count == 1 and self.get_selected_asset() and self.get_selected_asset().asset_type == "serial")
        self.delete_btn.setVisible(count > 0)

    def get_selected_asset(self):
        if len(self.selected_ids) != 1:
            return None
        aid = list(self.selected_ids)[0]
        for a in self.assets:
            if a.id == aid:
                return a
        return None

    def on_checkbox_changed(self, asset_id, state):
        if state == Qt.CheckState.Checked.value:
            self.selected_ids.add(asset_id)
        else:
            self.selected_ids.discard(asset_id)
        row = self._row_of_asset(asset_id)
        if row is not None:
            self._apply_row_selection(row)
        self.update_batch_actions()

    def _row_of_asset(self, asset_id):
        """资产 ID -> 当前表格行号（不在当前视图返回 None）"""
        for i, a in enumerate(getattr(self, "_table_assets", [])):
            if a.id == asset_id:
                return i
        return None

    def on_cell_entered(self, row, col):
        """鼠标悬停行变化：刷新新旧两行的复选框可见性"""
        if row == self.hovered_row:
            return
        old_row = self.hovered_row
        self.hovered_row = row
        self._set_row_checkbox_visible(old_row)
        self._set_row_checkbox_visible(row)

    def _set_row_checkbox_visible(self, row):
        """复选框可见性：已选中行常显，悬停行淡入显示，其余隐藏（不占位）"""
        if row is None or row < 0:
            return
        w = self.table.cellWidget(row, 0)
        if w is None:
            return
        cb = w.findChild(QCheckBox)
        if cb is None:
            return
        assets = getattr(self, "_table_assets", [])
        selected = row < len(assets) and assets[row].id in self.selected_ids
        should_show = selected or row == self.hovered_row
        if should_show and not cb.isVisible():
            cb.setVisible(True)
            self._fade_in_checkbox(cb)
        elif not should_show and cb.isVisible():
            # 先停掉进行中的淡入动画，避免动画结束后又把已隐藏控件绘出
            self._stop_cb_fade(cb)
            cb.setVisible(False)

    def _stop_cb_fade(self, cb):
        """安全停止复选框淡入动画并释放引用（C++ 对象可能已随控件销毁，须容错）"""
        anim = getattr(cb, "_fade_anim", None)
        if anim is not None:
            try:
                anim.stop()
                anim.deleteLater()
            except RuntimeError:
                pass  # 动画对象已被 Qt 销毁
            try:
                cb._fade_anim = None
            except RuntimeError:
                pass

    def _fade_in_checkbox(self, cb):
        """复选框悬停淡入（140ms 透明度 0→1，结束后移除效果恢复正常绘制）

        动画以复选框为父对象（控件销毁时随之销毁），不使用 DeleteWhenStopped，
        避免动画自然结束后 C++ 对象被删除而 Python 侧仍持有悬空引用。
        """
        self._stop_cb_fade(cb)
        eff = QGraphicsOpacityEffect(cb)
        cb.setGraphicsEffect(eff)
        anim = QPropertyAnimation(eff, b"opacity", cb)
        anim.setDuration(140)
        anim.setStartValue(0.0)
        anim.setEndValue(1.0)
        anim.setEasingCurve(QEasingCurve.Type.OutCubic)

        def _cleanup():
            try:
                cb.setGraphicsEffect(None)
                cb._fade_anim = None
            except RuntimeError:
                pass  # 表格重建后控件已销毁
            anim.deleteLater()

        anim.finished.connect(_cleanup)
        anim.start()
        cb._fade_anim = anim

    def on_cell_clicked(self, row, col):
        """单击行任意位置：切换该行选中状态（选中后行高亮）"""
        assets = getattr(self, "_table_assets", [])
        if row < 0 or row >= len(assets):
            return
        aid = assets[row].id
        if aid in self.selected_ids:
            self.selected_ids.discard(aid)
        else:
            self.selected_ids.add(aid)
        self._apply_row_selection(row)
        self.update_batch_actions()

    def _apply_row_selection(self, row):
        """同步某行的选中视觉：复选框勾选态 + 行高亮 + 复选框可见性（轻量，不重建表格）"""
        assets = getattr(self, "_table_assets", [])
        if row < 0 or row >= len(assets):
            return
        selected = assets[row].id in self.selected_ids
        w = self.table.cellWidget(row, 0)
        if w is not None:
            cb = w.findChild(QCheckBox)
            if cb is not None:
                cb.blockSignals(True)
                cb.setChecked(selected)
                cb.blockSignals(False)
        for col in range(self.table.columnCount()):
            item = self.table.item(row, col)
            if item is not None:
                item.setBackground(QColor("#eef2ff") if selected else QColor("transparent"))
        self._set_row_checkbox_visible(row)
        self.table.viewport().update()  # 触发高亮委托重绘该行

    def on_header_clicked(self, index):
        # 列索引 -> Asset 字段（第 0 列复选框、告警列与末列操作列不参与排序）
        fields = ["", "name", "asset_type", "account", "email", "total", "expire", "status", "", "note", "updated", ""]
        if 0 < index < len(fields) and fields[index]:
            field = fields[index]
            if self.sort_field == field:
                self.sort_asc = not self.sort_asc
            else:
                self.sort_field = field
                self.sort_asc = True
            self.sort_assets()
            # 表头显示排序方向箭头
            header = self.table.horizontalHeader()
            header.setSortIndicatorShown(True)
            header.setSortIndicator(
                index,
                Qt.SortOrder.AscendingOrder if self.sort_asc else Qt.SortOrder.DescendingOrder,
            )
            self.update_table()

    def sort_assets(self):
        """按当前字段排序；空值恒排末尾，数字与文本分开比较避免类型错误"""

        def key_func(a):
            val = getattr(a, self.sort_field, None)
            if val is None or val == "":
                return (1, 0, 0.0) if self.sort_asc else (-1, 0, 0.0)  # 空值始终沉底
            if isinstance(val, (int, float)):
                return (0, 0, float(val))
            return (0, 1, str(val).lower())

        self.assets.sort(key=key_func, reverse=not self.sort_asc)

    def _build_type_buttons(self):
        """按模板引擎重建侧栏类型筛选按钮（全部 + 分组内置模板 + 自定义模板）"""
        # 清空旧按钮与分组标题：脱离布局后控件仍挂在父级且可见，
        # 必须 hide + setParent(None) + deleteLater 彻底移除，避免残留
        while self.type_buttons_layout.count():
            item = self.type_buttons_layout.takeAt(0)
            w = item.widget()
            if w is not None:
                w.hide()
                w.setParent(None)
                w.deleteLater()
        self.type_buttons = {}

        def add_btn(key, icon, label, color):
            btn = SidebarButton(icon, label, 0, active=key == self.current_type,
                                active_style="type", color=color)
            btn.clicked.connect(lambda checked, k=key: self.on_type_filter(k))
            self.type_buttons[key] = btn
            self.type_buttons_layout.addWidget(btn)

        def add_group(title):
            lbl = QLabel(title)
            lbl.setStyleSheet(
                "font-size: 11px; font-weight: 700; color: #94a3b8;"
                "letter-spacing: 0.5px; padding: 12px 16px 4px 16px;"
            )
            self.type_buttons_layout.addWidget(lbl)

        add_btn("all", "📁", "全部资产", "#6366f1")
        builtin = {tpl["key"]: tpl for tpl in ordered_templates(self.custom_templates)}
        for group_title, keys in TEMPLATE_GROUPS:
            add_group(group_title)
            for key in keys:
                tpl = builtin[key]
                add_btn(tpl["key"], tpl["icon"], tpl["label"], tpl["color"])
        custom = [tpl for tpl in ordered_templates(self.custom_templates) if not tpl.get("builtin")]
        if custom:
            add_group("自定义")
            for tpl in custom:
                add_btn(tpl["key"], tpl["icon"], tpl["label"], tpl.get("color", "#0d9488"))
        # 当前筛选类型已不存在（如自定义模板被删）时回退到全部
        if self.current_type not in self.type_buttons:
            self.current_type = "all"
        for k, btn in self.type_buttons.items():
            btn.set_active(k == self.current_type)

    def on_type_filter(self, key):
        self.current_type = key
        for k, btn in self.type_buttons.items():
            btn.set_active(k == key)
        self.selected_ids.clear()
        titles = {"all": "全部资产", "serial": "序列号资产", "password": "密码资产"}
        subtitles = {
            "all": "管理您的序列号、密码与数字资产",
            "serial": "管理软件许可证与激活码",
            "password": "管理账号密码与访问凭证",
        }
        if key in titles:
            title, subtitle = titles[key], subtitles[key]
        else:
            tpl = get_template(key, self.custom_templates)
            title, subtitle = f"{tpl['label']}资产", f"管理{tpl['label']}类数字资产"
        self.page_title.setText(title)
        self.page_subtitle.setText(subtitle)
        self._update_column_visibility()
        self.update_stats()
        self.update_table()

    def _update_column_visibility(self):
        """上下文感知列：绑定邮箱仅对序列号/密码有意义；使用/数量仅序列号有数据"""
        self.table.setColumnHidden(4, self.current_type not in ("all", "serial", "password"))
        self.table.setColumnHidden(5, self.current_type not in ("all", "serial"))
        # 主字段列名跟随类型：序列号视图称「序列号」，密码视图称「账号」，其余/全部视图用通用名
        account_header = {"serial": "序列号", "password": "账号"}.get(self.current_type, "账号/序列号")
        self.table.horizontalHeaderItem(3).setText(account_header)

    def on_search(self, text):
        self.search_text = text
        self.selected_ids.clear()
        self.update_table()

    def eventFilter(self, obj, event):
        """捕获表格视口事件：右键抬起弹复制菜单，左键双击进入编辑，移出复位悬停行"""
        if obj is self.table.viewport():
            if event.type() == QEvent.Type.Leave:
                old_row = self.hovered_row
                self.hovered_row = -1
                self._set_row_checkbox_visible(old_row)
            # 右键菜单
            if (
                event.type() == QEvent.Type.MouseButtonRelease
                and event.button() == Qt.MouseButton.RightButton
            ):
                self.on_table_context_menu(event.pos())
                return True
            # 双击编辑（通过事件过滤器处理，跨平台稳定，
            # 不依赖各平台对 cellDoubleClicked 信号的合成差异）
            if (
                event.type() == QEvent.Type.MouseButtonDblClick
                and event.button() == Qt.MouseButton.LeftButton
            ):
                index = self.table.indexAt(event.pos())
                if index.isValid():
                    self.on_row_double_clicked(index.row(), index.column())
                return True
        return super().eventFilter(obj, event)

    def on_table_context_menu(self, pos):
        """右键复制：
        - 账号/序列号列：序列号类型复制序列号，密码类型复制密码
        - 有效期列：可复制有效期；密码类型还可复制网址
        """
        row = self.table.rowAt(pos.y())
        col = self.table.columnAt(pos.x())
        if row < 0 or col not in (1, 3, 4, 6):
            return
        assets = getattr(self, "_table_assets", [])
        if row >= len(assets):
            return
        asset = assets[row]

        actions = []  # [(菜单文本, 复制内容)]
        if col == 1:
            if asset.name:
                actions.append(("📋 复制名称", asset.name))
        elif col == 3:
            if asset.asset_type == "password":
                if asset.account:
                    actions.append(("📋 复制账号", asset.account))
                if asset.password:
                    actions.append(("📋 复制密码", asset.password))
            elif asset.asset_type == "serial":
                if asset.account:
                    actions.append(("📋 复制序列号", asset.account))
            else:
                # 模板引擎类型：主字段按模板字段名复制，密码类字段单独提供复制
                tpl = get_template(asset.asset_type, self.custom_templates)
                acc_spec = (tpl.get("columns") or {}).get("account", "")
                acc_key = acc_spec.split(",")[0].strip() if acc_spec else ""
                acc_label = next(
                    (f["label"] for f in tpl.get("fields", []) if f["key"] == acc_key),
                    "内容",
                )
                if asset.account and asset.account != "-":
                    actions.append((f"📋 复制{acc_label}", asset.account))
                for fdef in tpl.get("fields", []):
                    if (
                        fdef.get("type") == FIELD_PASSWORD
                        and fdef["key"] != acc_key
                        and asset.extra
                        and asset.extra.get(fdef["key"])
                    ):
                        actions.append((f"🔐 复制{fdef['label']}", asset.extra[fdef["key"]]))
        elif col == 4:
            if asset.email:
                actions.append(("📋 复制邮箱", asset.email))
        elif col == 6:
            # 有效期始终可复制：未设置时复制页面上显示的"永久"
            actions.append(("📅 复制有效期", asset.expire or "永久"))
            if asset.asset_type == "password" and asset.url:
                actions.append(("🔗 复制网址", asset.url))

        if not actions:
            self.show_toast("❌ 该记录无可复制内容")
            return

        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
                padding: 4px;
            }
            QMenu::item {
                padding: 10px 14px;
                font-size: 13px;
                color: #475569;
                border-radius: 6px;
            }
            QMenu::item:selected {
                background: #f8fafc;
            }
        """)
        for label, text in actions:
            act = QAction(label, self)
            act.triggered.connect(lambda checked=False, t=text: self.copy_to_clipboard(t))
            menu.addAction(act)
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def copy_to_clipboard(self, text):
        QApplication.clipboard().setText(text)
        self.show_toast("📋 已复制到剪贴板")

    def on_toggle_plaintext(self, checked):
        """切换账号/序列号 明文 <-> 掩码（默认明文）"""
        self.show_plaintext = checked
        self.plaintext_btn.setText("👁 明文显示" if checked else "🙈 掩码显示")
        self.update_table()
        self.show_toast("👁 已切换为明文显示" if checked else "🙈 已切换为掩码显示")

    def open_add_dialog(self, asset_type):
        dialog = AssetDialog(self, mode="add", asset_type=asset_type, custom_templates=self.custom_templates)
        dialog.saved.connect(self.on_asset_saved)
        dialog.exec()

    def _asset_by_id(self, aid):
        for a in self.assets:
            if a.id == aid:
                return a
        return None

    def on_edit(self):
        asset = self.get_selected_asset()
        if asset:
            self.on_edit_asset(asset.id)

    def on_edit_asset(self, aid):
        """按 ID 打开编辑对话框（双击行 / 操作列 ✏️ 共用）"""
        asset = self._asset_by_id(aid)
        if not asset:
            return
        dialog = AssetDialog(self, mode="edit", asset_type=asset.asset_type, asset=asset, custom_templates=self.custom_templates)
        dialog.saved.connect(self.on_asset_updated)
        dialog.exec()

    def on_copy_asset(self, aid):
        """操作列 📋：复制该记录的主凭据（密码类型复制密码，其余复制主字段）"""
        asset = self._asset_by_id(aid)
        if not asset:
            return
        if asset.asset_type == "password" and asset.password:
            text = asset.password
        else:
            text = asset.account or ""
        if text and text != "-":
            self.copy_to_clipboard(text)
        else:
            self.show_toast("❌ 该记录无可复制内容")

    def on_delete_asset(self, aid):
        """操作列 🗑️：删除该条记录（带确认）"""
        dialog = DeleteConfirmDialog(self, 1)
        dialog.confirmed.connect(lambda: self._delete_assets({aid}))
        dialog.exec()

    def on_row_double_clicked(self, row, col):
        """双击行任意位置：直接编辑该条记录"""
        assets = getattr(self, "_table_assets", [])
        if 0 <= row < len(assets):
            self.on_edit_asset(assets[row].id)

    def on_asset_saved(self, asset):
        asset.id = self.db.add_asset(asset)
        self.assets.append(asset)
        self.load_data()
        self.show_toast("✅ 新增成功")

    def on_asset_updated(self, asset):
        self.db.update_asset(asset)
        for i, a in enumerate(self.assets):
            if a.id == asset.id:
                self.assets[i] = asset
                break
        self.selected_ids.clear()
        self.load_data()
        self.show_toast("✅ 修改已保存")

    def on_delete(self):
        count = len(self.selected_ids)
        dialog = DeleteConfirmDialog(self, count)
        dialog.confirmed.connect(self.confirm_delete)
        dialog.exec()

    def confirm_delete(self):
        self._delete_assets(set(self.selected_ids))

    def _delete_assets(self, ids):
        """删除指定 ID 集合（批量删除与操作列单条删除共用）"""
        for aid in ids:
            self.db.delete_asset(aid)
        self.selected_ids.difference_update(ids)
        self.load_data()
        self.show_toast("🗑️ 已移入回收站")

    def on_toggle_archive(self, asset_id):
        """归档 / 取消归档：归档为手动停用，不计入活跃资产，也不再产生告警"""
        asset = next((a for a in self.assets if a.id == asset_id), None)
        if not asset:
            return
        if asset.status == "archived":
            asset.status = "normal"  # 先复位，再由 calculate_status 按最新数据重算
            asset.status = calculate_status(asset)
            toast = "↩️ 已取消归档，恢复为活跃资产"
        else:
            asset.status = "archived"
            toast = "📦 已归档，不再计入活跃资产"
        asset.updated = format_now()
        self.db.update_asset(asset)
        for i, a in enumerate(self.assets):
            if a.id == asset.id:
                self.assets[i] = asset
                break
        self.selected_ids.discard(asset_id)
        self.load_data()
        self.show_toast(toast)

    def on_use_once(self):
        asset = self.get_selected_asset()
        if not asset or asset.asset_type != "serial":
            return
        if (asset.remain or 0) <= 0:
            self.show_toast("❌ 剩余次数为 0，无法使用")
            return
        asset.used = (asset.used or 0) + 1
        asset.remain = max(0, (asset.total or 0) - asset.used)
        asset.status = calculate_status(asset)
        asset.updated = format_now()
        self.db.update_asset(asset)
        for i, a in enumerate(self.assets):
            if a.id == asset.id:
                self.assets[i] = asset
                break
        self.selected_ids.clear()
        self.load_data()
        self.show_toast(f"📌 已使用一次，剩余 {asset.remain} 次")

    def open_recycle_bin(self):
        dialog = RecycleBinDialog(self, self.recycle_bin, custom_templates=self.custom_templates)
        dialog.restored.connect(self.on_restore)
        dialog.permanently_deleted.connect(self.on_permanent_delete)
        dialog.cleared.connect(self.on_clear_recycle_bin)
        dialog.restored_all.connect(self.on_restore_all)
        dialog.exec()
        self.load_data()

    def on_clear_recycle_bin(self):
        self.db.clear_recycle_bin()
        self.show_toast("🧹 回收站已清空")

    def on_restore_all(self):
        self.db.restore_all()
        self.show_toast("↩️ 已全部恢复")

    def on_restore(self, idx):
        if 0 <= idx < len(self.recycle_bin):
            asset = self.recycle_bin[idx]
            self.db.restore_asset(asset.id)
            self.show_toast("↩️ 已恢复")

    def on_permanent_delete(self, idx):
        if 0 <= idx < len(self.recycle_bin):
            asset = self.recycle_bin[idx]
            self.db.permanent_delete(asset.id)
            self.show_toast("🗑️ 已彻底删除")

    def create_export_menu(self):
        """导出菜单：JSON 备份 / Excel 表格"""
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
                padding: 4px;
            }
            QMenu::item {
                padding: 10px 14px;
                font-size: 13px;
                color: #475569;
                border-radius: 6px;
            }
            QMenu::item:selected {
                background: #f8fafc;
            }
            QMenu::separator {
                height: 1px;
                background: #f1f5f9;
                margin: 4px 10px;
            }
        """)
        json_action = QAction("📄 导出 JSON 备份", self)
        json_action.triggered.connect(self.export_data)
        menu.addAction(json_action)
        menu.addSeparator()
        csv_action = QAction("📑 导出 CSV (.csv)", self)
        csv_action.triggered.connect(self.export_csv)
        menu.addAction(csv_action)
        excel_action = QAction("📊 导出 Excel (.xlsx)", self)
        excel_action.triggered.connect(self.export_excel)
        menu.addAction(excel_action)
        return menu

    def export_csv(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "导出 CSV", "asset_vault_export.csv", "CSV 文件 (*.csv)"
        )
        if not path:
            return
        try:
            self._write_csv(path)
            self.show_toast("📑 已导出 CSV 文件")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出 CSV 失败: {e}")

    def _write_csv(self, path):
        """写出 CSV（utf-8-sig 带 BOM，Excel 打开中文不乱码）"""
        import csv
        headers = [
            "所属", "名称", "类型", "账号/序列号", "密码", "绑定邮箱",
            "总数量", "已用次数", "剩余次数", "有效期/网址",
            "状态", "备注", "更新时间", "扩展字段", "标签", "删除时间",
        ]
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for a in self.db.get_assets():
                writer.writerow(["资产"] + self._asset_excel_row(a, self.custom_templates) + [""])
            for a in self.db.get_recycle_bin():
                writer.writerow(["回收站"] + self._asset_excel_row(a, self.custom_templates) + [a.deleted_at or ""])

    def export_data(self):
        data = self.db.export_data()
        path, _ = QFileDialog.getSaveFileName(self, "导出备份", "asset_vault_backup.json", "JSON (*.json)")
        if path:
            with open(path, "w", encoding="utf-8") as f:
                f.write(data)
            self.show_toast("📤 已导出 JSON 备份")

    @staticmethod
    def _excel_width(text) -> int:
        """估算列宽（中文按 2 个字符宽计）"""
        s = "" if text is None else str(text)
        return sum(2 if ord(ch) > 127 else 1 for ch in s)

    def _fill_excel_sheet(self, ws, headers, rows):
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="6366F1")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin", color="E2E8F0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for row in rows:
            ws.append(row)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")
        for idx, col in enumerate(ws.columns, 1):
            width = max((self._excel_width(c.value) for c in col), default=8)
            ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 4, 10), 46)
        ws.freeze_panes = "A2"

    @staticmethod
    def _asset_excel_row(a, custom_templates=None):
        tc = get_type_config(a.asset_type, custom_templates)
        st = STATUS_CONFIG.get(a.status, STATUS_CONFIG["normal"])
        if a.asset_type == "serial":
            expire_or_url = a.expire or "永久"
        elif a.asset_type == "password":
            expire_or_url = a.url or "-"
        else:
            expire_or_url = a.expire or "-"
        # 模板引擎扩展字段：渲染为「字段名: 值」对
        extra_str = ""
        if a.extra:
            tpl = get_template(a.asset_type, custom_templates)
            labels = {f["key"]: f["label"] for f in tpl.get("fields", [])}
            extra_str = "; ".join(f"{labels.get(k, k)}: {v}" for k, v in a.extra.items())
        return [
            a.name,
            tc["label"],
            a.account,                       # 明文导出
            a.password or "-",
            a.email or "-",
            a.total if a.total is not None else "-",
            a.used if a.used is not None else "-",
            a.remain if a.remain is not None else "-",
            expire_or_url,
            st["label"],
            a.note or "",
            a.updated,
            extra_str,
            a.tags or "",
        ]

    def export_excel(self):
        try:
            from openpyxl import Workbook
        except ImportError:
            QMessageBox.warning(
                self, "缺少依赖",
                "导出 Excel 需要 openpyxl，请先安装：\npip install openpyxl"
            )
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "导出 Excel", "asset_vault_export.xlsx", "Excel 文件 (*.xlsx)"
        )
        if not path:
            return
        try:
            wb = Workbook()
            headers = [
                "名称", "类型", "账号/序列号", "密码", "绑定邮箱",
                "总数量", "已用次数", "剩余次数", "有效期/网址",
                "状态", "备注", "更新时间", "扩展字段", "标签",
            ]
            ws1 = wb.active
            ws1.title = "资产"
            self._fill_excel_sheet(
                ws1, headers,
                [self._asset_excel_row(a, self.custom_templates) for a in self.db.get_assets()],
            )
            ws2 = wb.create_sheet("回收站")
            self._fill_excel_sheet(
                ws2, headers + ["删除时间"],
                [
                    self._asset_excel_row(a, self.custom_templates) + [a.deleted_at or ""]
                    for a in self.db.get_recycle_bin()
                ],
            )
            wb.save(path)
            self.show_toast("📊 已导出 Excel 文件")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出 Excel 失败: {e}")

    # ===== 设置 =====
    def create_settings_menu(self):
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
                padding: 4px;
            }
            QMenu::item {
                padding: 10px 14px;
                font-size: 13px;
                color: #475569;
                border-radius: 6px;
            }
            QMenu::item:selected {
                background: #f8fafc;
            }
            QMenu::separator {
                height: 1px;
                background: #f1f5f9;
                margin: 4px 10px;
            }
        """)
        change_pw_action = QAction("🔑 修改主密码", self)
        change_pw_action.triggered.connect(self.change_master_password)
        menu.addAction(change_pw_action)
        tpl_action = QAction("🧩 资产模板管理", self)
        tpl_action.triggered.connect(self.open_template_manager)
        menu.addAction(tpl_action)
        menu.addSeparator()
        about_action = QAction("ℹ️ 关于 AssetVault", self)
        about_action.triggered.connect(self.show_about)
        menu.addAction(about_action)
        return menu

    def open_template_manager(self):
        """打开资产模板管理（自定义模板、字段类型、到期提醒规则）"""
        dialog = TemplateManagerDialog(self, self.db)
        dialog.changed.connect(self.on_templates_changed)
        dialog.exec()

    def on_templates_changed(self):
        """模板变更后：刷新缓存、侧栏筛选按钮与新增菜单"""
        self.custom_templates = self.db.get_custom_templates()
        self._build_type_buttons()
        self.add_btn.setMenu(self.create_add_menu())
        self.load_data()

    def open_settings(self):
        menu = self.create_settings_menu()
        menu.exec(self.settings_btn.mapToGlobal(QPoint(0, self.settings_btn.height() + 4)))

    def change_master_password(self):
        dialog = ChangeMasterPasswordDialog(self)
        dialog.submitted.connect(self.on_master_password_submitted)
        dialog.exec()

    def on_master_password_submitted(self, old_pw: str, new_pw: str):
        if self.db.change_master_password(old_pw, new_pw):
            QMessageBox.information(self, "成功", "主密码修改成功，下次启动请使用新密码解锁")
        else:
            QMessageBox.warning(self, "失败", "当前主密码不正确，修改失败")

    def show_about(self):
        QMessageBox.information(
            self, "关于 AssetVault",
            "AssetVault v2.0\n个人数字资产管理器\n\n本地 AES-256-GCM 加密存储 · 离线运行"
        )

    def import_data(self):
        path, _ = QFileDialog.getOpenFileName(self, "导入备份", "", "JSON (*.json)")
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = f.read()
            self.db.import_data(data)
            self.load_data()
            self.show_toast("📥 导入成功")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入失败: {e}")

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if self.toast.isVisible():
            self.toast.move(
                self.width() - self.toast.width() - 24,
                self.height() - self.toast.height() - 24
            )
